# run_demo.py
"""
Pipeline 快速測試工具

使用方式：直接改 MODE 和測試內容，python run_demo.py 就能跑
"""
import spacy
import tempfile
import os
import shutil
import re
import csv
from pathlib import Path

from pipeline import MaskingPipeline, DialogueTurn
from pseudonym import PseudonymTracker


# ══════════════════════════════════════════════════════════════
# 設定區（直接改這裡）
# ══════════════════════════════════════════════════════════════

MODE = "csv"

# ── CSV 模式（讀 Excel，輸出 CSV）─────────────────────────
CSV_PATH   = "/raid2/Kee/Harry借用/Liang借用/04_語音資料庫/測試檔/部分_1.xlsx"       # 輸入 Excel 路徑
CSV_OUTPUT = "/raid2/Kee/Harry借用/Liang借用/04_語音資料庫/測試檔/test.csv"     # 輸出 CSV 路徑
CSV_COLUMN = "文本"                              # 對話內容的欄位名稱
MAX_ROWS   = 20                                       # 測幾筆，None 跑全部

# ── 單句模式 ──────────────────────────────────────────────
SINGLE_TEXT = """我叫黃鼎量，身分證字號是A123456789，信用卡卡號是1234567898765432，住址是台北市大安區仁愛路四段1號5樓
"""

# ── 多輪對話模式 ──────────────────────────────────────────
DIALOGUE = [
    ("AGENT",    "請問您的大名是？"),
    ("USER", "我叫王小明"),
    ("AGENT",    "請問您的身分證字號？"),
    ("USER", "A123456789"),
    ("AGENT",    "電話號碼幾號呢？"),
    ("USER", "0912345678"),
    ("AGENT",    "信用卡卡號是多少？"),
    ("USER", "卡號1234567890123456"),
    ("AGENT",    "需要轉帳多少？"),
    ("USER", "我要轉帳50000元到帳號12345678901234"),
]

# ── Pipeline 設定 ─────────────────────────────────────────
THRESHOLD   = 0.50
CKIP_MODEL  = "/raid2/model/ckip/bert-base-chinese-ner"
CKIP_DEVICE = -1   # -1=CPU, 0=GPU
ENABLE_CKIP = True

# ══════════════════════════════════════════════════════════════


def init_pipeline():
    tmpdir = tempfile.mkdtemp()
    blank_path = os.path.join(tmpdir, "blank_zh")
    spacy.blank("zh").to_disk(blank_path)

    pipeline = MaskingPipeline(
        log_path=None,
        score_threshold=THRESHOLD,
        spacy_model=blank_path,
        enable_ckip=ENABLE_CKIP,
        ckip_model=CKIP_MODEL,
        ckip_device=CKIP_DEVICE,
    )
    return pipeline, tmpdir


def run_single(pipeline):
    print("=" * 60)
    print("  單句測試")
    print("=" * 60)

    r = pipeline.mask(SINGLE_TEXT, session_id="single_test")

    print(f"\n  原始：{SINGLE_TEXT}")
    if r.normalized_text != SINGLE_TEXT:
        print(f"  正規：{r.normalized_text}")
    print(f"  脫敏：{r.masked_text}")
    if r.entities_found:
        for e in r.entities_found:
            print(f"  　→ {e.entity_type}: {r.normalized_text[e.start:e.end]} (score={e.score:.2f})")
    else:
        print("  　→ (無偵測到實體)")
    print(f"  可用度：{r.usability_tag}")


def run_dialogue(pipeline):
    print("=" * 60)
    print("  多輪對話測試")
    print("=" * 60)

    tracker = PseudonymTracker(session_id="dialogue_test")

    for i, (speaker, text) in enumerate(DIALOGUE, 1):
        r = pipeline.mask(
            text=text,
            session_id="dialogue_test",
            speaker=speaker,
            diarization_available=True,
            tracker=tracker,
        )
        print(f"\n  [Turn {i:02d}] {speaker}")
        print(f"    原始：{text}")
        if r.normalized_text != text:
            print(f"    正規：{r.normalized_text}")
        print(f"    脫敏：{r.masked_text}")
        if r.entities_found:
            for e in r.entities_found:
                print(f"    　→ {e.entity_type}: {r.normalized_text[e.start:e.end]}")

def parse_dialogue_text(raw_text: str):
    """
    解析對話文字，格式：
    '坐席:    0.24    4.18    讓您久等...\n客戶:    1.40    1.47    好\n'
    """
    turns = []
    pattern = re.compile(r"(坐席|客戶):\s+[\d.]+\s+[\d.]+\s+(.+)")
    for line in str(raw_text).strip().split("\n"):
        line = line.strip()
        if not line:
            continue
        m = pattern.match(line)
        if m:
            role = "AGENT" if m.group(1) == "坐席" else "CUSTOMER"
            text = m.group(2).strip()
            if text:
                turns.append((role, text))
    return turns

def run_csv(pipeline):
    import openpyxl

    print("=" * 60)
    print(f"  Excel → CSV 脫敏：{CSV_PATH}")
    print("=" * 60)

    wb = openpyxl.load_workbook(CSV_PATH, data_only=True)
    ws = wb.active

    # 讀取 header
    headers = [cell.value for cell in ws[1]]
    col_idx = headers.index(CSV_COLUMN) if CSV_COLUMN in headers else None
    if col_idx is None:
        print(f"[錯誤] 找不到欄位 '{CSV_COLUMN}'，現有欄位：{headers}")
        return

    # 輸出 CSV
    out_path = Path(CSV_OUTPUT)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    out_headers = headers + ["dialogue_masked", "entities_found"]

    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(out_headers)

        total = 0
        total_entities = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
            if MAX_ROWS and row_idx > MAX_ROWS:
                break

            raw = row[col_idx] if col_idx < len(row) else ""
            row_list = list(row)

            if not raw or not str(raw).strip():
                writer.writerow(row_list + ["", ""])
                total += 1
                continue

            turns = parse_dialogue_text(str(raw))
            if not turns:
                writer.writerow(row_list + [raw, ""])
                total += 1
                continue

            tracker = PseudonymTracker(session_id=f"row_{row_idx}")
            masked_lines = []
            all_entities = []

            for speaker, text in turns:
                r = pipeline.mask(
                    text=text,
                    session_id=f"row_{row_idx}",
                    speaker=speaker,
                    diarization_available=True,
                    tracker=tracker,
                )
                role_label = "坐席" if speaker == "AGENT" else "客戶"
                masked_lines.append(f"{role_label}: {r.masked_text}")

                for e in r.entities_found:
                    all_entities.append(
                        f"{e.entity_type}({r.normalized_text[e.start:e.end]})"
                    )

            masked_text = "\n".join(masked_lines)
            entities_str = ", ".join(all_entities) if all_entities else ""

            writer.writerow(row_list + [masked_text, entities_str])
            total += 1
            total_entities += len(all_entities)

            if all_entities:
                print(f"  [Row {row_idx}] {len(all_entities)} 個實體：{', '.join(all_entities[:5])}{'...' if len(all_entities) > 5 else ''}")

    print(f"\n  輸出 → {out_path}")
    print(f"  共 {total} 筆，偵測 {total_entities} 個實體")


def main():
    pipeline, tmpdir = init_pipeline()

    try:
        if MODE == "single":
            run_single(pipeline)
        elif MODE == "dialogue":
            run_dialogue(pipeline)
        elif MODE == "csv":
            run_csv(pipeline)
        else:
            print(f"[錯誤] 未知 MODE: {MODE}")
    finally:
        pipeline.close()
        shutil.rmtree(tmpdir, ignore_errors=True)

    print("\n" + "=" * 60)
    print("  完成！")
    print("=" * 60)


if __name__ == "__main__":
    main()